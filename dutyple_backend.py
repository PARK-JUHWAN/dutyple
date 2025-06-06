# -*- coding: utf-8 -*-

# 라이브러리 임포트
import uuid
import os
import json
import pandas as pd
import random
import calendar
import holidays
import string
import openpyxl
from collections import defaultdict
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from flask import Flask, request, send_file, jsonify, g
import logging
from threading import Thread
import time
import datetime

# --- 기본 설정 ---
UPLOAD_FOLDER = 'uploads'
RESULT_FOLDER = 'results'
LOG_FILE = "log.txt"
FILE_RETENTION_HOURS = 24  # 파일 보관 시간 (24시간)

# 폴더 생성
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(RESULT_FOLDER, exist_ok=True)

# Flask 앱 초기화
app = Flask(__name__)

# --- 로깅 및 파일 관리 함수 ---
def write_log(text):
    """로그 파일에 텍스트를 기록합니다."""
    with open(LOG_FILE, "a", encoding="utf-8") as f:
        f.write(text + "\n")
    print(text) # 콘솔에도 출력

def clear_log():
    """서버 시작 시 로그 파일을 초기화합니다."""
    if os.path.exists(LOG_FILE):
        os.remove(LOG_FILE)

def cleanup_files():
    """주기적으로 오래된 파일을 정리하는 함수 (백그라운드 스레드에서 실행)"""
    while True:
        try:
            write_log(f"[SYSTEM] {FILE_RETENTION_HOURS}시간 이상된 파일 정리 시작...")
            now = time.time()
            retention_period_seconds = FILE_RETENTION_HOURS * 3600

            for folder in [UPLOAD_FOLDER, RESULT_FOLDER]:
                for filename in os.listdir(folder):
                    file_path = os.path.join(folder, filename)
                    if os.path.isfile(file_path):
                        file_mod_time = os.path.getmtime(file_path)
                        if (now - file_mod_time) > retention_period_seconds:
                            os.remove(file_path)
                            write_log(f" - 오래된 파일 삭제: {file_path}")
            
            # 1시간에 한 번씩 실행
            time.sleep(3600)
        except Exception as e:
            write_log(f"[ERROR] 파일 정리 중 오류 발생: {e}")
            time.sleep(3600) # 오류 발생 시에도 1시간 대기


# --- 핵심 근무표 생성 로직 ---
def run_dutyple(uid, input_path, output_path, config):
    """
    프론트엔드에서 전달받은 설정(config)을 기반으로 근무표 생성 알고리즘을 실행합니다.
    모든 관련 로직이 이 함수 내에서 캡슐화되어 실행됩니다.
    """
    try:
        write_log("Dutyple 생성을 시작합니다.")
        write_log(f"설정값: {json.dumps(config, ensure_ascii=False, indent=2)}")
        
        # --- 설정값 변수 할당 ---
        nurse_count = config["nurse_count"]
        year = config["year"]
        month = config["month"]
        N_count_nurse = config["N_count_nurse"]
        weekday_D = config["weekday_D"]
        weekday_E = config["weekday_E"]
        weekday_N = config["weekday_N"]
        holiday_D = config["holiday_D"]
        holiday_E = config["holiday_E"]
        holiday_N = config["holiday_N"]

        weekday_W = weekday_D + weekday_E
        holiday_W = holiday_D + holiday_E
        
        # --- 1. 엑셀 파일 처리 및 이름 매핑 ---
        write_log("1/5 - 엑셀 파일 처리 및 이름 매핑 중...")
        wb = openpyxl.load_workbook(input_path)
        ws = wb.active
        name_map = {}
        for i in range(2, 2 + nurse_count):
            name = ws.cell(row=i, column=1).value
            if name is None:
                break
            name_map[name] = string.ascii_uppercase[i - 2]
            ws.cell(row=i, column=1).value = string.ascii_uppercase[i - 2]
        wb.save(input_path)
        
        # --- 2. 변수 초기화 ---
        write_log("2/5 - 변수 초기화 중...")
        nurse = list(string.ascii_uppercase[:nurse_count])
        duty = ["N", "W", "X"]
        weight = {"N": 0, "W": 1, "X": 2}
        _, num_day = calendar.monthrange(year, month)

        kr_holidays = holidays.CountryHoliday('KR', years=year)
        month_holiday = [day.day for day in kr_holidays if day.month == month]
        month_end = {day for day in range(1, num_day + 1) if calendar.weekday(year, month, day) in [5, 6]}
        daily_end = sorted(list(month_end | set(month_holiday)))
        daily = list(range(1, num_day + 1))

        daily_wallet = {}
        for day in daily:
            if day in daily_end:
                daily_wallet[day] = {'D': holiday_D, 'E': holiday_E, 'N': holiday_N, 'W': holiday_W, 'X': holiday_W}
            else:
                daily_wallet[day] = {'D': weekday_D, 'E': weekday_E, 'N': weekday_N, 'W': weekday_W, 'X': weekday_W}

        nurse_wallet = {}
        for worker in nurse:
            X_count_nurse = len(daily_end)
            W_count_nurse = len(daily) - N_count_nurse - X_count_nurse
            nurse_wallet[f"{worker}_wallet"] = {"N": N_count_nurse, "X": X_count_nurse, "W": W_count_nurse}

        df_origin = pd.read_excel(input_path, index_col=0)
        df = pd.DataFrame(index=df_origin.index, columns=[-2, -1, 0] + daily)
        
        # --- 3. 사전 근무 배정 (Prefer) ---
        write_log("3/5 - 사전 근무 배정(Prefer) 적용 중...")
        def prefer(nurse, day, duty_char):
            if f"{nurse}_wallet" not in nurse_wallet: return
            if nurse_wallet[f"{nurse}_wallet"][duty_char] > 0:
                nurse_wallet[f"{nurse}_wallet"][duty_char] -= 1
                if day in daily_wallet and daily_wallet[day][duty_char] > 0:
                    daily_wallet[day][duty_char] -= 1
                    df.loc[nurse, day] = duty_char
                else:
                    nurse_wallet[f"{nurse}_wallet"][duty_char] += 1
        
        for nurse_name in df_origin.index:
            for day_val in df_origin.columns:
                if isinstance(day_val, int):
                    raw_val = df_origin.at[nurse_name, day_val]
                    if pd.notna(raw_val):
                        duty_val = str(raw_val).strip().upper()
                        if duty_val in ["D", "E"]:
                            prefer(nurse_name, day_val, "W")
                        elif duty_val in ["N", "X"]:
                            prefer(nurse_name, day_val, duty_val)
        
        # --- 4. 메인 로직 실행 (z-rule, ND-rule 등) ---
        write_log("4/5 - 메인 알고리즘 실행 중...")
        z_rules = {0:["X"], 2:["X","W","N"], 6:["N"], 7:["W","X","N"], 8:["W","X","N"], 9:["N","X"], 11:["X","W","N"], 12:["N"], 13:["X","N","W"], 14:["W","X","N"], 15:["N"], 16:["W","X","N"], 17:["W","X","N"], 18:["N","X"], 20:["W","X","N"], 21:["N"], 22:["W","X","N"], 23:["W","X","N"], 24:["N"], 25:["W","X","N"], 26:["W","N","X"]}

        def can_assign(nurse_name, day, duty_char):
            return nurse_wallet[f"{nurse_name}_wallet"][duty_char] > 0 and daily_wallet[day][duty_char] > 0

        max_global_attempt = 1000
        global_attempt = 0
        success_all = False

        while global_attempt < max_global_attempt and not success_all:
            global_attempt += 1
            write_log(f"   - 전체 시도 {global_attempt}회차 시작")
            df_backup = df.copy()
            daily_wallet_backup = {k: v.copy() for k, v in daily_wallet.items()}
            nurse_wallet_backup = {k: v.copy() for k, v in nurse_wallet.items()}
            
            success_all = True
            for today in daily:
                if pd.isna(df.loc[:, today]).all():
                    nurse_list = list(df.index)
                    random.shuffle(nurse_list)
                    for nurse_name in nurse_list:
                        if pd.isna(df.loc[nurse_name, today]):
                            # Ensure past days are in df.columns before accessing
                            a = df.loc[nurse_name, today - 3] if today - 3 in df.columns else 'X'
                            b = df.loc[nurse_name, today - 2] if today - 2 in df.columns else 'X'
                            c = df.loc[nurse_name, today - 1] if today - 1 in df.columns else 'X'
                            z = 9 * weight.get(a, 2) + 3 * weight.get(b, 2) + weight.get(c, 2)
                            duty_candidates = z_rules.get(z, ["W", "N", "X"])
                            random.shuffle(duty_candidates)
                            for duty_char in duty_candidates:
                                if can_assign(nurse_name, today, duty_char):
                                    prefer(nurse_name, today, duty_char)
                                    break
                if daily_wallet[today]['N']!=0 or daily_wallet[today]['W']!=0:
                    success_all = False
                    break
            
            if success_all and any(w['N'] > 1 for w in nurse_wallet.values()):
                 success_all = False
            
            if not success_all:
                df = df_backup
                daily_wallet = daily_wallet_backup
                nurse_wallet = nurse_wallet_backup

        if not success_all:
             raise Exception("근무 배정에 실패했습니다. (최대 시도 횟수 초과)")
        
        # --- 5. 최종 처리 및 엑셀 저장 ---
        write_log("5/5 - 최종 결과 처리 및 파일 저장 중...")
        w_map = defaultdict(list)
        for nurse_name in df.index:
            for day_val in df.columns:
                if isinstance(day_val, int) and df.loc[nurse_name, day_val] == 'W':
                    w_map[day_val].append(nurse_name)
        
        for day_val, nurses in w_map.items():
            random.shuffle(nurses)
            d_count = daily_wallet_backup[day_val]['D']
            for i, nurse_name in enumerate(nurses):
                prev_duty = df.loc[nurse_name, day_val - 1] if day_val - 1 in df.columns else None
                if i < d_count and prev_duty != 'E':
                    df.loc[nurse_name, day_val] = 'D'
                else:
                    df.loc[nurse_name, day_val] = 'E'
        
        # [로직 추가] ED룰 위반 찾기
        ed_violations = []
        for nurse_name in df.index:
            for day_val in daily:
                if day_val > 1: # 첫째 날은 검사할 필요 없음
                    prev_day = day_val - 1
                    if df.loc[nurse_name, prev_day] == 'E' and df.loc[nurse_name, day_val] == 'D':
                        ed_violations.append((nurse_name, day_val))
        write_log(f"E->D 금지 위반 감지: {len(ed_violations)}건")

        summary_cols = ["D", "E", "N", "X"]
        for col in summary_cols: df[col] = 0
        for nurse_name in df.index:
            for col in summary_cols:
                df.loc[nurse_name, col] = (df.loc[nurse_name, daily] == col).sum()
        
        # --- 엑셀 파일 쓰기 및 스타일링 ---
        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name="dutyple")
            wb_result = writer.book
            ws_result = wb_result["dutyple"]
            
            # 기본 스타일링 루프
            for col_idx, day in enumerate(df.columns, start=2):
                col_letter = get_column_letter(col_idx)
                for row_idx, nurse_name in enumerate(df.index, start=2):
                    cell = ws_result.cell(row=row_idx, column=col_idx)
                    val = df.loc[nurse_name, day]

                    # [로직 추가] 사전 입력(Prefer) 근무는 빨간색 글씨로
                    is_prefer = (
                        day in df_origin.columns and
                        nurse_name in df_origin.index and
                        pd.notna(df_origin.loc[nurse_name, day])
                    )
                    
                    # 배경색 (공휴일/주말)
                    if isinstance(day, int) and day in daily_end:
                        cell.fill = PatternFill("solid", fgColor="FFE4B5")
                    
                    # 글씨색
                    if is_prefer:
                        cell.font = Font(color="FF0000", bold=True)
                    elif val == "N":
                        cell.font = Font(color="CC9900")

            # [로직 추가] ED룰 위반 셀에 하늘색 배경 적용
            ed_fail_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
            for nurse_name, day in ed_violations:
                if day in df.columns and nurse_name in df.index:
                    row = df.index.get_loc(nurse_name) + 2
                    col = df.columns.get_loc(day) + 2
                    ws_result.cell(row=row, column=col).fill = ed_fail_fill

        final_wb = openpyxl.load_workbook(output_path)
        ws_final = final_wb["dutyple"]
        rev_map = {v: k for k, v in name_map.items()}
        for row_idx in range(2, 2 + len(rev_map)):
            alias = ws_final.cell(row=row_idx, column=1).value
            if alias in rev_map:
                ws_final.cell(row=row_idx, column=1).value = rev_map[alias]
        final_wb.save(output_path)

        write_log(f"배정 성공! 엑셀 저장 완료 → {output_path}")

    except Exception as e:
        write_log(f"오류 발생: {str(e)}")
        raise e


# --- Flask API 엔드포인트 ---

@app.route('/upload', methods=['POST'])
def upload_file():
    """파일과 설정값을 업로드하고 임시 저장합니다."""
    clear_log()
    write_log("'/upload' 요청을 받았습니다.")
    
    if 'file' not in request.files:
        return jsonify({"error": "파일이 없습니다."}), 400
    file = request.files['file']
    if file.filename == '':
        return jsonify({"error": "파일이 선택되지 않았습니다."}), 400

    try:
        uid = uuid.uuid4().hex[:8]
        input_path = os.path.join(UPLOAD_FOLDER, f'{uid}.xlsx')
        file.save(input_path)
        write_log(f"파일 저장 완료: {input_path}")

        config = {
            "nurse_count": int(request.form['nurse_count']),
            "year": int(request.form['year']),
            "month": int(request.form['month']),
            "weekday_D": int(request.form['weekday_D']),
            "weekday_E": int(request.form['weekday_E']),
            "weekday_N": int(request.form['weekday_N']),
            "holiday_D": int(request.form['holiday_D']),
            "holiday_E": int(request.form['holiday_E']),
            "holiday_N": int(request.form['holiday_N']),
            "N_count_nurse": int(request.form['N_count_nurse'])
        }
        
        config_path = os.path.join(UPLOAD_FOLDER, f'config_{uid}.json')
        with open(config_path, 'w', encoding='utf-8') as f:
            json.dump(config, f)
        write_log(f"설정 저장 완료: {config_path}")
        
        return jsonify({"uuid": uid})
    except Exception as e:
        write_log(f"업로드 처리 중 오류: {str(e)}")
        return jsonify({"error": "서버 처리 중 오류가 발생했습니다."}), 500

@app.route('/generate', methods=['GET'])
def generate_schedule():
    """저장된 파일과 설정을 기반으로 근무표 생성을 시작합니다."""
    uid = request.args.get('uuid')
    if not uid:
        return jsonify({"error": "UUID가 필요합니다."}), 400

    write_log(f"'/generate' 요청을 받았습니다. (UUID: {uid})")
    
    input_path = os.path.join(UPLOAD_FOLDER, f'{uid}.xlsx')
    output_path = os.path.join(RESULT_FOLDER, f'result_{uid}.xlsx')
    config_path = os.path.join(UPLOAD_FOLDER, f'config_{uid}.json')

    if not os.path.exists(input_path) or not os.path.exists(config_path):
        write_log("오류: 요청한 UUID에 해당하는 파일이나 설정을 찾을 수 없습니다.")
        return jsonify({"error": "파일 또는 설정을 찾을 수 없습니다."}), 404

    with open(config_path, 'r', encoding='utf-8') as f:
        config = json.load(f)

    thread = Thread(target=run_dutyple, args=(uid, input_path, output_path, config))
    thread.start()
    
    return jsonify({"status": "generation_started", "uuid": uid}), 202

@app.route('/download-template', methods=['GET'])
def download_template():
    """미리 준비된 엑셀 템플릿 파일을 다운로드합니다."""
    template_path = 'dutyple_template.xlsx'
    if not os.path.exists(template_path):
        return jsonify({"error": "템플릿 파일을 찾을 수 없습니다."}), 404
    return send_file(template_path, as_attachment=True, download_name='dutyple_template.xlsx')

@app.route('/result/<uid>', methods=['GET'])
def get_result(uid):
    """지정된 uid에 해당하는 결과 파일을 다운로드합니다."""
    path = os.path.join(RESULT_FOLDER, f"result_{uid}.xlsx")
    if not os.path.exists(path):
        return jsonify({"error": "결과 파일을 찾을 수 없거나 아직 생성 중입니다."}), 404
    return send_file(path, as_attachment=True, download_name=f'dutyple_result_{uid}.xlsx')


@app.route('/log', methods=['GET'])
def get_log():
    """생성 과정의 로그를 반환합니다."""
    if not os.path.exists(LOG_FILE):
        return '', 204 # No Content
    with open(LOG_FILE, "r", encoding="utf-8") as f:
        return f.read()

# --- 서버 실행 ---
if __name__ == '__main__':
    # 서버 실행 시 로그 파일 초기화
    clear_log()
    
    # 템플릿 파일이 없다면 기본 템플릿 생성
    if not os.path.exists('dutyple_template.xlsx'):
        wb_template = openpyxl.Workbook()
        ws_template = wb_template.active
        ws_template.title = "dutyple"
        ws_template['A1'] = "간호사 이름"
        ws_template['B1'] = -2
        ws_template['C1'] = -1
        ws_template['D1'] = 0
        ws_template['A2'] = "김간호"
        ws_template['A3'] = "이간호"
        ws_template['A4'] = "박간호"
        wb_template.save('dutyple_template.xlsx')
        print("기본 템플릿 파일 'dutyple_template.xlsx'을 생성했습니다.")
    
    # 오래된 파일 정리를 위한 백그라운드 스레드 시작
    cleanup_thread = Thread(target=cleanup_files, daemon=True)
    cleanup_thread.start()
    
    app.run(host='0.0.0.0', port=10000, debug=True)


# 추가 작업 / 제미나이 여기는 제가 메모장으로 쓴 부분이기에 무시해주세요
# 저연차끼리의 근무
# 나이트 갯수가 오버 될 수 있기에, 자동으로 +1 여유
# NightKeep, DayEveKeep, 개인별 근무 제한 설정
# 요일별로도 다르게 설정
# '3일' 대신 원하는 기간(예: 5일)을 설정할 수 있습니다