import os
import sys
import pandas as pd
import random
import calendar
import holidays
import string
import openpyxl
from collections import defaultdict
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

desktop = os.path.join(os.path.expanduser("~"), "Desktop")
input_path = os.path.join(desktop, "dutyple.xlsx")
output_path = os.path.join(desktop, "★dutyple★.xlsx")

wb = openpyxl.load_workbook(input_path)
ws = wb.active
name_map = {}

for i in range(2, 2+26):
    name = ws.cell(row=i, column=1).value
    if name is None:
        break
    name_map[name] = string.ascii_uppercase[i - 2]
    ws.cell(row=i, column=1).value = string.ascii_uppercase[i - 2]

wb.save(input_path)


print("6-1) 간호사 수는 몇 명인가요?")
nurse_count = int(input().replace(" ", ""))

print("6-2) 몇 년도 근무표를 짜시겠어요?")
year_input = input().replace(" ", "")
year = int(year_input)
if year < 100:
    year += 2000

print("6-3) 몇 월 근무표를 짜시겠어요?")
month_input = input().replace(" ", "")
month = int(month_input.lstrip("0")) if month_input != "0" else 0

print("6-4) 간호사 1인당 월간 나이트 개수는 몇 개인가요?")
n_input = input("   기본값: ").replace(" ", "")
N_count_nurse = int(n_input) if n_input == "0" else int(n_input) + 1

def parse_duty_input(msg):
    parts = input(msg).replace(" ", "").split("/")
    if len(parts) != 3 or not all(p.isdigit() for p in parts):
        print("숫자/숫자/숫자 형태로 입력해주세요")
        sys.exit(1)
    return map(int, parts)

print("6-5) 평일 기준, 데이/이브닝/나이트 인원을 입력해주세요")
weekday_D, weekday_E, weekday_N = parse_duty_input("   예: 2/3/2 → ")

print("6-6) 주말&공휴일 기준, 데이/이브닝/나이트 인원을 입력해주세요")
holiday_D, holiday_E, holiday_N = parse_duty_input("   예: 1/1/2 → ")

print("엑셀 파일을 다운받고, 이전달 3일치의 근무를 입력하세요")
print("클릭하시면 근무표가 바탕화면에 생성됩니다")

weekday_W = weekday_D + weekday_E
holiday_W = holiday_D + holiday_E

if nurse_count > 26:
    sys.exit(1)

nurse = list(string.ascii_uppercase[:nurse_count])
duty = ["N", "W", "X"]
weight = {"N": 0, "W": 1, "X": 2}

_, num_day = calendar.monthrange(year, month)

def get(year, month):
    month_holiday = [day.day for day in holidays.CountryHoliday('KR', years=year) if day.month == month]
    month_end = {day for day in range(1, num_day + 1) if calendar.weekday(year, month, day) in [5, 6]}
    valid_holend = month_end - set(month_holiday)
    return month_holiday, month_end, valid_holend

month_holiday, month_end, valid_holend = get(year, month)
daily_all = [-2, -1, 0] + list(range(1, num_day + 1))
daily = list(range(1, num_day + 1))
daily_end = sorted(month_end | set(month_holiday))
daily_week = [day for day in daily if day not in daily_end]

daily_wallet = {}
for day in daily:
    if day in daily_end:
        daily_wallet[day] = {'N': holiday_N, 'W': holiday_W, 'X': holiday_W}  # X도 같은 값 유지
    else:
        daily_wallet[day] = {'N': weekday_N, 'W': weekday_W, 'X': weekday_W}  # X도 같은 값 유지

nurse_wallet = {}
for worker in nurse:
    X_count_nurse = len(daily_end) + 1
    W_count_nurse = len(daily) - N_count_nurse - X_count_nurse + 4
    nurse_wallet[f"{worker}_wallet"] = {"N": N_count_nurse, "X": X_count_nurse, "W": W_count_nurse}

cell_header = daily_all
cell_index = nurse
df = pd.DataFrame(index=cell_index, columns=cell_header)

def prefer(nurse, day, duty):
    if f"{nurse}_wallet" in nurse_wallet:
        if duty == "N" and nurse_wallet[f"{nurse}_wallet"]["N"] > 0:
            nurse_wallet[f"{nurse}_wallet"]["N"] -= 1
        elif duty == "X" and nurse_wallet[f"{nurse}_wallet"]["X"] > 0:
            nurse_wallet[f"{nurse}_wallet"]["X"] -= 1
        elif duty == "W" and nurse_wallet[f"{nurse}_wallet"]["W"] > 0:
            nurse_wallet[f"{nurse}_wallet"]["W"] -= 1
    if day in daily_wallet:
        if daily_wallet[day][duty] > 0:
            daily_wallet[day][duty] -= 1
    df.loc[nurse, day] = duty

# 입력 엑셀 불러오기
df_origin = pd.read_excel(input_path, index_col=0)
origin_cols = list(df_origin.columns)
extra_cols = [-2, -1, 0]
final_cols = extra_cols + [c for c in origin_cols if c not in extra_cols]
df = pd.DataFrame(index=df_origin.index, columns=final_cols)

for col in [-2, -1, 0]:
    if col in df_origin.columns:
        for nurse_name in df.index:
            val = df_origin.loc[nurse_name, col]
            if val in ["D", "E"]:
                df.loc[nurse_name, col] = "W"
            elif val in ["N", "X"]:
                df.loc[nurse_name, col] = val
            else:
                df.loc[nurse_name, col] = None

for nurse_name in df_origin.index:
    for day in df_origin.columns:
        if isinstance(day, int) and day > 0:
            try:
                raw_val = df_origin.at[nurse_name, day]
            except KeyError:
                continue  # 열 이름이 잘못되었거나 존재하지 않는 경우 건너뜀
            if pd.notna(raw_val):
                duty = str(raw_val).strip().upper()
                if duty in ["D", "E"]:
                    prefer(nurse_name, day, "W")
                elif duty in ["N", "X"]:
                    prefer(nurse_name, day, duty)


z_rules = {
    0:  ["X"],             # N-N-N
    2:  ["X", "W", "N"],   # N-N-X
    6:  ["N"],             # N-X-N
    7:  ["W", "X", "N"],   # N-X-W
    8:  ["W", "X", "N"],   # N-X-X
    9:  ["N", "X"],        # W-N-N
    11: ["X", "W", "N"],   # W-N-X
    12: ["N"],             # W-W-N
    13: ["X", "N", "W"],   # W-W-W
    14: ["W", "X", "N"],   # W-W-X
    15: ["N"],             # W-X-N
    16: ["W", "X", "N"],   # W-X-W
    17: ["W", "X", "N"],   # W-X-X
    18: ["N", "X"],        # X-N-N
    20: ["W", "X", "N"],   # X-N-X
    21: ["N"],             # X-W-N
    22: ["W", "X", "N"],   # X-W-W
    23: ["W", "X", "N"],   # X-W-X
    24: ["N"],             # X-X-N
    25: ["W", "X", "N"],   # X-X-W
    26: ["W", "N", "X"],        # X-X-X
}

def can_assign(nurse_name, day, duty):
    if nurse_wallet[f"{nurse_name}_wallet"][duty] <= 0:
        return False
    if daily_wallet[day][duty] <= 0:
        return False
    return True
max_global_attempt = 1000000
global_attempt = 0
success_all = False
day_range = [day for day in df.columns if isinstance(day, int) and day > 0]
day_range.sort()

# ND_rule
while global_attempt < max_global_attempt:
    df_backup = df.copy()
    daily_wallet_backup = {day: daily_wallet[day].copy() for day in daily_wallet}
    nurse_wallet_backup = {n: nurse_wallet[n].copy() for n in nurse_wallet}
    print(f"\n🔁 전체 재시도 {global_attempt + 1}회차 시작")

    success_all = True
    for today in day_range:
        past_days = [today - 3, today - 2, today - 1]
        if any(p not in df.columns for p in past_days):
            continue
        max_attempt = 100
        success = False

        for attempt in range(max_attempt):
            df_day_backup = df[today].copy()
            day_wallet_backup = daily_wallet[today].copy()
            nurse_wallet_inner_backup = {n: nurse_wallet[n].copy() for n in nurse_wallet}
            assigned = []
            nurse_list = list(df.index)
            random.shuffle(nurse_list)

            for nurse_name in nurse_list:
                a = df.loc[nurse_name, past_days[0]]
                b = df.loc[nurse_name, past_days[1]]
                c = df.loc[nurse_name, past_days[2]]
                if pd.isna(a) or pd.isna(b) or pd.isna(c):
                    continue

                if today in df_origin.columns and nurse_name in df_origin.index:
                    if pd.notna(df_origin.loc[nurse_name, today]):
                        duty_prefer = df.loc[nurse_name, today]
                        z = 9 * weight.get(a, 0) + 3 * weight.get(b, 0) + weight.get(c, 0)
                        duty_candidates = z_rules.get(z, ["W", "N", "X"])

                        if duty_prefer in duty_candidates:
                            assigned.append(nurse_name)
                            continue
                        else:
                            adjusted = False
                            for a_try in ["N", "W", "X"]:
                                for b_try in ["N", "W", "X"]:
                                    for c_try in ["N", "W", "X"]:
                                        z_try = 9 * weight[a_try] + 3 * weight[b_try] + weight[c_try]
                                        allowed = z_rules.get(z_try, [])
                                        if duty_prefer in allowed:
                                            can_use = True
                                            for d, val in zip(past_days, [a_try, b_try, c_try]):
                                                if not can_assign(nurse_name, d, val):
                                                    can_use = False
                                                    break
                                            if can_use:
                                                for d, val in zip(past_days, [a_try, b_try, c_try]):
                                                    prefer(nurse_name, d, val)
                                                adjusted = True
                                                break
                                    if adjusted: break
                                if adjusted: break
                            if not adjusted:
                                success = False
                                success_all = False
                                break
                        continue
                z = 9 * weight.get(a, 0) + 3 * weight.get(b, 0) + weight.get(c, 0)
                duty_candidates = z_rules.get(z, ["W", "N", "X"])
                for duty in duty_candidates:
                    if can_assign(nurse_name, today, duty):
                        prefer(nurse_name, today, duty)
                        assigned.append(nurse_name)
                        break
            if len(assigned) == len(nurse) and all(v == 0 for v in daily_wallet[today].values()):
                print(f"✅ Day {today} 배정 성공 (시도 {attempt+1}회)")
                success = True
                break
            else:
                df[today] = df_day_backup
                daily_wallet[today] = day_wallet_backup
                for n in nurse_wallet:
                    nurse_wallet[n] = nurse_wallet_inner_backup[n].copy()
        if not success:
            print(f"🟥 Day {today} 배정 실패 (최대 {max_attempt}회 시도)")
            success_all = False
            break

# N2_rule
    if success_all:
        if any(remain["N"] >= 2 for remain in nurse_wallet.values()):
        # if any(remain["N"] >= 1 for remain in nurse_wallet.values()):
            df = df_backup.copy()
            daily_wallet = {day: daily_wallet_backup[day].copy() for day in daily_wallet}
            nurse_wallet = {n: nurse_wallet_backup[n].copy() for n in nurse_wallet}
            global_attempt += 1
            continue
        break
    else:
        df = df_backup.copy()
        daily_wallet = {day: daily_wallet_backup[day].copy() for day in daily_wallet}
        nurse_wallet = {n: nurse_wallet_backup[n].copy() for n in nurse_wallet}
        global_attempt += 1
if not success_all:
   pass

# ED_rule
for col in [-2, -1, 0]:
    if col in df_origin.columns:
        for nurse_name in df.index:
            original_val = df_origin.loc[nurse_name, col]
            if df.loc[nurse_name, col] == "W" and original_val in ["D", "E"]:
                df.loc[nurse_name, col] = original_val
for nurse_name in df.index:
    for day in df_origin.columns:
        if isinstance(day, int) and day > 0:
            original_val = df_origin.loc[nurse_name, day]
            if df.loc[nurse_name, day] == "W" and original_val in ["D", "E"]:
                df.loc[nurse_name, day] = original_val
w_map = defaultdict(list)
for nurse_name in df.index:
    for day in df.columns:
        if isinstance(day, int) and df.loc[nurse_name, day] == "W":
            w_map[day].append(nurse_name)
for day, nurse_list in w_map.items():
    random.shuffle(nurse_list)
    if day in daily_end:
        d_quota = holiday_D
        e_quota = holiday_E
    else:
        d_quota = weekday_D
        e_quota = weekday_E
    for i, nurse_name in enumerate(nurse_list):
        prev_day = day - 1
        prev_is_E = False
        if isinstance(prev_day, int) and prev_day in df.columns:
            prev_val = df.loc[nurse_name, prev_day]
            if isinstance(prev_val, str) and prev_val == "E":
                prev_is_E = True
        if i < d_quota:
            if prev_is_E:
                df.loc[nurse_name, day] = "E"
            else:
                df.loc[nurse_name, day] = "D"
        else:
            df.loc[nurse_name, day] = "E"
def ED_rule(df, df_backup, daily_wallet, daily_wallet_backup, nurse_wallet, nurse_wallet_backup):
    for nurse in df.index:
        for day in df.columns:
            if not isinstance(day, int): continue
            if df.loc[nurse, day] == "E":
                next_day = day + 1
                if next_day in df.columns and df.loc[nurse, next_day] == "D":
                    for col in df.columns:
                        df[col] = df_backup[col]
                    daily_wallet.clear()
                    daily_wallet.update({day: daily_wallet_backup[day].copy() for day in daily_wallet})
                    nurse_wallet.clear()
                    nurse_wallet.update({n: nurse_wallet_backup[n].copy() for n in nurse_wallet})
                    return False
    return True
def ED_rule(df):
    ED_fail = []
    for nurse in df.index:
        for day in df.columns:
            if not isinstance(day, int): continue
            if df.loc[nurse, day] == "E":
                next_day = day + 1
                if next_day in df.columns and df.loc[nurse, next_day] == "D":
                    ED_fail.append((nurse, day))
    return ED_fail
ED_fail = ED_rule(df)

# excel
summary_cols = ["D", "E", "N", "X"]
for col in summary_cols:
    df[col] = 0
valid_days = [col for col in df.columns if isinstance(col, int) and col > 0]
for nurse_name in df.index:
    count_d = (df.loc[nurse_name, valid_days] == "D").sum()
    count_e = (df.loc[nurse_name, valid_days] == "E").sum()
    count_n = (df.loc[nurse_name, valid_days] == "N").sum()
    count_x = (df.loc[nurse_name, valid_days] == "X").sum()
    df.loc[nurse_name, "D"] = count_d
    df.loc[nurse_name, "E"] = count_e
    df.loc[nurse_name, "N"] = count_n
    df.loc[nurse_name, "X"] = count_x

with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
    df.to_excel(writer, sheet_name="dutyple")
    wb = writer.book
    ws = wb["dutyple"]
    for col_idx, day in enumerate(df.columns, start=2):
        col_letter = get_column_letter(col_idx)
        for row_idx, nurse in enumerate(df.index, start=2):
            cell = ws.cell(row=row_idx, column=col_idx)
            val = df.loc[nurse, day]
            is_prefer = (
                day in df_origin.columns and
                nurse in df_origin.index and
                pd.notna(df_origin.loc[nurse, day]))
            if day in [-2, -1, 0]:
                cell.fill = PatternFill("solid", fgColor="DDDDDD")
            elif isinstance(day, int) and day in daily_end:
                cell.fill = PatternFill("solid", fgColor="FFE4B5")
            if is_prefer:
                cell.font = Font(color="FF0000")
            elif val == "N":
                cell.font = Font(color="CC9900")
    ED_failfill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
    for nurse, day in ED_fail:
        if day in df.columns:  
            r = df.index.get_loc(nurse) + 2
            c = list(df.columns).index(day) + 2
            ws.cell(row=r, column=c).fill = ED_failfill

final_wb = openpyxl.load_workbook(output_path)
ws_result = final_wb["dutyple"]

rev_map = {v: k for k, v in name_map.items()}

for row_idx in range(2, 2 + len(rev_map)):
    alias = ws_result.cell(row=row_idx, column=1).value
    if alias in rev_map:
        ws_result.cell(row=row_idx, column=1).value = rev_map[alias]

final_wb.save(output_path)
